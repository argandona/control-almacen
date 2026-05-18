import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction

from django.db.models import Sum, Q

from .models import (
    Usuario, Rol, Empresa, Camion, Material, StockCamion,
    UploadConsumo, Consumo, DetalleConsumo,
    Pedido, DetallePedido,
    Devolucion, DetalleDevolucion,
    Inventario, DetalleInventario,
    SST, UsuarioCamion,
)
from .security import (
    verificar_clave, hashear_clave, migrar_clave_si_legacy,
    esta_bloqueado, registrar_intento_fallido, limpiar_intentos, intentos_restantes,
)


# ── Auth helpers ──────────────────────────────────────────────────────────────

ROLES_PERMITIDOS = (Rol.LIQUIDADOR, Rol.ENCARGADO_ALMACEN, Rol.SUPERADMIN)


def _usuario_session(request):
    uid = request.session.get('portal_uid')
    if not uid:
        return None
    try:
        return Usuario.objects.select_related('rol', 'empresa').get(pk=uid, activo=True)
    except Usuario.DoesNotExist:
        return None


def _login_required(view_fn):
    def wrapper(request, *args, **kwargs):
        if not _usuario_session(request):
            return redirect('portal_login')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def _superadmin_required(view_fn):
    def wrapper(request, *args, **kwargs):
        usuario = _usuario_session(request)
        if not usuario:
            return redirect('portal_login')
        if usuario.rol_id != Rol.SUPERADMIN:
            messages.error(request, 'Acceso restringido al SuperAdmin.')
            return redirect('portal_consumos')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


# ── Vistas ────────────────────────────────────────────────────────────────────

def portal_login(request):
    usuario_activo = _usuario_session(request)
    if usuario_activo:
        if usuario_activo.rol_id == Rol.SUPERADMIN:
            return redirect('portal_usuarios')
        return redirect('portal_consumos')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        clave = request.POST.get('clave', '').strip()

        if esta_bloqueado(email):
            messages.error(request, 'Demasiados intentos fallidos. Intenta en 15 minutos.')
            return render(request, 'portal/login.html')

        try:
            usuario = Usuario.objects.select_related('rol').get(email=email, activo=True)
        except Usuario.DoesNotExist:
            registrar_intento_fallido(email)
            messages.error(request, 'Correo o contraseña incorrectos.')
            return render(request, 'portal/login.html')

        if not verificar_clave(clave, usuario.clave):
            restantes = intentos_restantes(email) - 1
            registrar_intento_fallido(email)
            msg = 'Correo o contraseña incorrectos.'
            if 0 <= restantes <= 2:
                msg += f' Te quedan {restantes} intento(s) antes del bloqueo.'
            messages.error(request, msg)
            return render(request, 'portal/login.html')

        if usuario.rol_id not in ROLES_PERMITIDOS:
            messages.error(request, 'Tu rol no tiene acceso al portal.')
            return render(request, 'portal/login.html')

        limpiar_intentos(email)
        migrar_clave_si_legacy(usuario, clave)

        from django.utils import timezone
        usuario.ultimo_acceso = timezone.now()
        usuario.save(update_fields=['ultimo_acceso'])

        request.session['portal_uid'] = usuario.id_usuario
        request.session.set_expiry(28800)
        if usuario.rol_id == Rol.SUPERADMIN:
            return redirect('portal_usuarios')
        return redirect('portal_consumos')

    return render(request, 'portal/login.html')


def portal_logout(request):
    request.session.flush()
    return redirect('portal_login')


@_login_required
def portal_consumos(request):
    usuario  = _usuario_session(request)
    resultado = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Selecciona un archivo Excel (.xlsx).')
        elif not archivo.name.endswith('.xlsx'):
            messages.error(request, 'Solo se aceptan archivos .xlsx')
        else:
            resultado = _procesar_y_aprobar(request, archivo, usuario)

    return render(request, 'portal/consumos.html', {
        'usuario':   usuario,
        'resultado': resultado,
    })


def portal_plantilla(request):
    """Descarga plantilla Excel con todas las matrículas de materiales como columnas."""
    materiales = list(
        Material.objects.order_by('matricula').values_list('matricula', flat=True)
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consumos'

    header_font  = Font(bold=True, color='FFFFFF', size=10)
    header_fill  = PatternFill('solid', fgColor='1D6A3A')
    mat_fill     = PatternFill('solid', fgColor='155724')
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin         = Side(style='thin', color='CCCCCC')
    borde        = Border(left=thin, right=thin, top=thin, bottom=thin)

    fixed = ['SST', 'SUMINISTRO', 'FECHA EJECUCION\n(dd/mm/aaaa)', 'TECNICO_EMPLEADO\n(Apellido Nombre)']
    all_headers = fixed + list(materiales)

    ws.row_dimensions[1].height = 36
    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill if col <= 4 else mat_fill
        cell.alignment = center
        cell.border    = borde
        ws.column_dimensions[cell.column_letter].width = max(13, len(h.split('\n')[0]) + 4)

    # Fila de ejemplo
    ejemplo = ['SST-001', 'S-12345', '15/05/2026', 'GARCIA LOPEZ JUAN'] + [0] * len(materiales)
    for col, val in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.border    = borde
        cell.alignment = Alignment(horizontal='center')

    # Nota al pie
    ws.cell(row=4, column=1,
            value='⚠ Todas las filas deben pertenecer al mismo código SST. '
                  'Ingresar 0 si no se usó el material.')

    ws.freeze_panes = 'E2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_consumos.xlsx"'
    wb.save(response)
    return response


# ── Lógica de procesamiento ───────────────────────────────────────────────────

def _procesar_y_aprobar(request, archivo, usuario_upload):
    """
    Flujo en tres pasos (todo en memoria antes de tocar la BD):
      1. Parsear Excel → detectar errores de formato/técnico/fecha.
      2. Verificar stock para TODOS los materiales → mostrar todos los faltantes.
      3. Solo si paso 1 y 2 sin errores → guardar en BD y descontar stock.
    """
    try:
        wb   = openpyxl.load_workbook(archivo, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
    except Exception as exc:
        messages.error(request, f'Error al leer el archivo: {exc}')
        return None

    if len(rows) < 2:
        messages.error(request, 'El archivo no contiene datos (solo la fila de encabezado).')
        return None

    header           = rows[0]
    matriculas_excel = [str(h).strip() for h in header[4:] if h is not None]

    materiales_map = {
        m.matricula: m
        for m in Material.objects.filter(matricula__in=matriculas_excel)
    }

    codigo_sst = str(rows[1][0]).strip() if rows[1][0] is not None else ''
    try:
        sst = SST.objects.get(codigo=codigo_sst)
    except SST.DoesNotExist:
        messages.error(request, f'El código SST "{codigo_sst}" no existe en el sistema.')
        return None

    # Verificar si ya existe un upload aprobado para este SST
    upload_existente = UploadConsumo.objects.filter(sst=sst).first()
    if upload_existente and upload_existente.estado == 'aprobado':
        return {'ya_aprobado': True, 'sst': codigo_sst, 'upload': upload_existente}

    # ── Paso 1: Parsear Excel en memoria ─────────────────────────────────────
    # plan = lista de dicts con todos los datos ya validados
    errores_parseo = []
    plan = []  # [{tecnico, camion, suministro, fecha, materiales: [(material, qty), ...]}]

    for num_fila, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        suministro  = str(row[1]).strip() if row[1] is not None else ''
        fecha_raw   = row[2]
        tecnico_str = str(row[3]).strip().upper() if row[3] is not None else ''
        cantidades  = row[4:]

        if isinstance(fecha_raw, (datetime.date, datetime.datetime)):
            fecha = fecha_raw if isinstance(fecha_raw, datetime.date) else fecha_raw.date()
        else:
            try:
                fecha = datetime.datetime.strptime(str(fecha_raw).strip(), '%d/%m/%Y').date()
            except ValueError:
                errores_parseo.append(f'Fila {num_fila}: fecha inválida "{fecha_raw}" (usa dd/mm/aaaa).')
                continue

        qs_tec = Usuario.objects.filter(nombre__icontains=tecnico_str)
        if not qs_tec.exists():
            errores_parseo.append(f'Fila {num_fila}: técnico "{tecnico_str}" no encontrado.')
            continue
        if qs_tec.count() > 1:
            nombres = ', '.join(qs_tec.values_list('nombre', flat=True)[:3])
            errores_parseo.append(
                f'Fila {num_fila}: "{tecnico_str}" coincide con varios usuarios ({nombres}…).')
            continue
        tecnico = qs_tec.first()

        camion = UsuarioCamion.camion_activo_de_usuario(tecnico, fecha)
        if not camion:
            errores_parseo.append(
                f'Fila {num_fila}: "{tecnico.nombre}" sin camión asignado el {fecha:%d/%m/%Y}.')
            continue

        items_fila = []
        for i, cantidad in enumerate(cantidades):
            if i >= len(matriculas_excel):
                break
            try:
                qty = int(cantidad) if cantidad not in (None, '') else 0
            except (TypeError, ValueError):
                qty = 0
            if qty == 0:
                continue
            material = materiales_map.get(matriculas_excel[i])
            if not material:
                errores_parseo.append(
                    f'Fila {num_fila}: matrícula "{matriculas_excel[i]}" no existe en el sistema.')
                continue
            items_fila.append((material, qty, camion))

        plan.append({
            'tecnico':    tecnico,
            'camion':     camion,
            'suministro': suministro,
            'fecha':      fecha,
            'items':      items_fila,
        })

    if errores_parseo:
        return {
            'aprobado': False,
            'sst':      codigo_sst,
            'archivo':  archivo.name,
            'errores':  errores_parseo,
        }

    # ── Paso 2: Acumular totales a descontar por (camion, material) ───────────
    totales = {}
    for fila in plan:
        for material, qty, camion in fila['items']:
            key = (camion.pk, material.pk)
            if key not in totales:
                totales[key] = {'qty': 0, 'material': material, 'camion': camion}
            totales[key]['qty'] += qty

    # ── Paso 3: Guardar en BD y descontar stock (puede quedar negativo) ───────
    with transaction.atomic():
        upload, _ = UploadConsumo.objects.get_or_create(
            sst=sst,
            defaults={
                'usuario':        usuario_upload,
                'nombre_archivo': archivo.name,
                'estado':         'pendiente',
            },
        )

        creados = 0
        for fila in plan:
            consumo, nuevo = Consumo.objects.get_or_create(
                upload=upload,
                usuario_consume=fila['tecnico'],
                camion=fila['camion'],
                suministro=fila['suministro'],
                fecha=fila['fecha'],
            )
            for material, qty, _ in fila['items']:
                DetalleConsumo.objects.get_or_create(
                    consumo=consumo,
                    material=material,
                    defaults={'cantidad': qty},
                )
            if nuevo:
                creados += 1

        # Descontar stock (puede quedar negativo si no hay registro previo)
        for info in totales.values():
            stock, _ = StockCamion.objects.get_or_create(
                camion=info['camion'], material=info['material'],
                defaults={'cantidad': 0},
            )
            stock.cantidad -= info['qty']
            stock.save(update_fields=['cantidad'])

        upload.estado = 'aprobado'
        upload.save()

    return {
        'aprobado': True,
        'sst':      codigo_sst,
        'archivo':  archivo.name,
        'creados':  creados,
        'errores':  [],
        'upload':   upload,
    }


# ── Pedidos / Devoluciones / Matriz (Enc. Almacén + SuperAdmin) ──────────────

def _almacen_required(view_fn):
    def wrapper(request, *args, **kwargs):
        usuario = _usuario_session(request)
        if not usuario:
            return redirect('portal_login')
        if usuario.rol_id not in (Rol.ENCARGADO_ALMACEN, Rol.SUPERADMIN):
            messages.error(request, 'Acceso restringido al Encargado de Almacén.')
            return redirect('portal_consumos')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def _camiones_empresa(usuario):
    """Devuelve QS de camiones visibles para el usuario."""
    if usuario.empresa_id:
        return Camion.objects.filter(empresa_id=usuario.empresa_id, activo=True).order_by('placa')
    return Camion.objects.filter(activo=True).order_by('placa')


@_almacen_required
def portal_pedidos(request):
    usuario  = _usuario_session(request)
    camiones = _camiones_empresa(usuario)

    camion_id  = request.GET.get('camion')
    estado_fil = request.GET.get('estado', '')
    desde      = request.GET.get('desde', '')
    hasta      = request.GET.get('hasta', '')

    qs = (Pedido.objects
          .filter(camion__in=camiones)
          .select_related('camion', 'usuario', 'usuario_aprueba')
          .prefetch_related('detalles__material')
          .order_by('-fecha'))

    if camion_id:  qs = qs.filter(camion_id=camion_id)
    if estado_fil: qs = qs.filter(estado=estado_fil)
    if desde:      qs = qs.filter(fecha__gte=desde)
    if hasta:      qs = qs.filter(fecha__lte=hasta)

    return render(request, 'portal/pedidos.html', {
        'usuario':    usuario,
        'pedidos':    qs,
        'camiones':   camiones,
        'camion_id':  camion_id or '',
        'estado_fil': estado_fil,
        'desde':      desde,
        'hasta':      hasta,
    })


@_almacen_required
def portal_pedidos_excel(request):
    """Exporta los pedidos filtrados a Excel con una fila por detalle de material."""
    usuario  = _usuario_session(request)
    camiones = _camiones_empresa(usuario)

    camion_id  = request.GET.get('camion')
    estado_fil = request.GET.get('estado', '')
    desde      = request.GET.get('desde', '')
    hasta      = request.GET.get('hasta', '')

    qs = (Pedido.objects
          .filter(camion__in=camiones)
          .select_related('camion', 'usuario', 'usuario_aprueba')
          .prefetch_related('detalles__material')
          .order_by('-fecha'))

    if camion_id:  qs = qs.filter(camion_id=camion_id)
    if estado_fil: qs = qs.filter(estado=estado_fil)
    if desde:      qs = qs.filter(fecha__gte=desde)
    if hasta:      qs = qs.filter(fecha__lte=hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    # Estilos
    verde       = PatternFill('solid', fgColor='1D6A3A')
    hdr_font    = Font(bold=True, color='FFFFFF', size=11)
    alt_fill    = PatternFill('solid', fgColor='E8F5E9')
    center      = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        '# Pedido', 'Fecha', 'Estado', 'Camión',
        'Solicitado por', 'Aprobado / Despachado por', 'Fecha aprobación',
        'Matrícula', 'Material', 'Cant. solicitada', 'Cant. despachada',
    ]
    col_widths = [10, 14, 12, 12, 24, 28, 18, 14, 36, 16, 16]

    ws.append(headers)
    for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = hdr_font
        cell.fill      = verde
        cell.alignment = center
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    row_num = 2
    for pedido in qs:
        aprobado_por  = pedido.usuario_aprueba.nombre if pedido.usuario_aprueba else '—'
        fecha_aprob   = pedido.fecha_aprobacion.strftime('%d/%m/%Y') if pedido.fecha_aprobacion else '—'
        detalles      = list(pedido.detalles.all())

        for det in detalles:
            fila = [
                pedido.id_pedido,
                pedido.fecha.strftime('%d/%m/%Y') if pedido.fecha else '—',
                pedido.get_estado_display(),
                pedido.camion.placa,
                pedido.usuario.nombre,
                aprobado_por,
                fecha_aprob,
                det.material.matricula,
                det.material.descripcion,
                det.cantidad_solicitada,
                det.cantidad_aprobada if det.cantidad_aprobada else 0,
            ]
            ws.append(fila)
            fill = alt_fill if row_num % 2 == 0 else None
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = thin_border
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill
            row_num += 1

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Nombre del archivo con filtros aplicados
    sufijo = f'_{estado_fil}' if estado_fil else ''
    sufijo += f'_{desde}_{hasta}' if desde or hasta else ''
    filename = f'pedidos{sufijo}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@_almacen_required
def portal_devoluciones(request):
    usuario  = _usuario_session(request)
    camiones = _camiones_empresa(usuario)

    camion_id  = request.GET.get('camion')
    estado_fil = request.GET.get('estado', '')
    desde      = request.GET.get('desde', '')
    hasta      = request.GET.get('hasta', '')

    qs = (Devolucion.objects
          .filter(camion__in=camiones)
          .select_related('camion', 'usuario', 'usuario_aprueba')
          .prefetch_related('detalles__material')
          .order_by('-fecha'))

    if camion_id:  qs = qs.filter(camion_id=camion_id)
    if estado_fil: qs = qs.filter(estado=estado_fil)
    if desde:      qs = qs.filter(fecha__gte=desde)
    if hasta:      qs = qs.filter(fecha__lte=hasta)

    return render(request, 'portal/devoluciones.html', {
        'usuario':     usuario,
        'devoluciones': qs,
        'camiones':    camiones,
        'camion_id':   camion_id or '',
        'estado_fil':  estado_fil,
        'desde':       desde,
        'hasta':       hasta,
    })


@_almacen_required
def portal_matriz(request):
    usuario  = _usuario_session(request)
    camiones = _camiones_empresa(usuario)

    camion_id = request.GET.get('camion')
    desde     = request.GET.get('desde', '')
    hasta     = request.GET.get('hasta', '')

    camiones_sel = camiones.filter(pk=camion_id) if camion_id else camiones

    # ── Filtros de fecha para movimientos ────────────────────────────────────
    q_ped = Q(pedido__camion__in=camiones_sel, pedido__estado='aprobado')
    q_con = Q(consumo__camion__in=camiones_sel)
    q_dev = Q(devolucion__camion__in=camiones_sel, devolucion__estado='aprobado')
    if desde:
        q_ped &= Q(pedido__fecha__gte=desde)
        q_con &= Q(consumo__fecha__gte=desde)
        q_dev &= Q(devolucion__fecha__gte=desde)
    if hasta:
        q_ped &= Q(pedido__fecha__lte=hasta)
        q_con &= Q(consumo__fecha__lte=hasta)
        q_dev &= Q(devolucion__fecha__lte=hasta)

    # ── Aggregaciones por material ────────────────────────────────────────────
    pedidos_map = {
        r['material_id']: r['total']
        for r in DetallePedido.objects.filter(q_ped).values('material_id').annotate(total=Sum('cantidad_aprobada'))
    }
    consumos_map = {
        r['material_id']: r['total']
        for r in DetalleConsumo.objects.filter(q_con).values('material_id').annotate(total=Sum('cantidad'))
    }
    devoluciones_map = {
        r['material_id']: r['total']
        for r in DetalleDevolucion.objects.filter(q_dev).values('material_id').annotate(total=Sum('cantidad_aprobada'))
    }
    stocks_map = {
        r['material_id']: r['total']
        for r in StockCamion.objects.filter(camion__in=camiones_sel).values('material_id').annotate(total=Sum('cantidad'))
    }

    # ── Último inventario cerrado por camión → físico por material ────────────
    from django.db.models import Max
    last_inv_ids = list(
        Inventario.objects
        .filter(camion__in=camiones_sel, estado='cerrado')
        .values('camion_id')
        .annotate(last_id=Max('id_inventario'))
        .values_list('last_id', flat=True)
    )
    inventario_map = {
        r['material_id']: r['total']
        for r in DetalleInventario.objects
            .filter(inventario_id__in=last_inv_ids)
            .values('material_id')
            .annotate(total=Sum('cantidad_fisica'))
    }

    # ── Unir todos los materiales con movimiento ──────────────────────────────
    material_ids = (
        set(pedidos_map) | set(consumos_map) |
        set(devoluciones_map) | set(stocks_map) | set(inventario_map)
    )
    materiales = Material.objects.filter(pk__in=material_ids).order_by('matricula')

    filas = []
    for m in materiales:
        p   = pedidos_map.get(m.pk, 0) or 0
        c   = consumos_map.get(m.pk, 0) or 0
        d   = devoluciones_map.get(m.pk, 0) or 0
        s   = stocks_map.get(m.pk, 0) or 0
        inv = inventario_map.get(m.pk)          # None = sin inventario registrado
        filas.append({
            'material':          m,
            'pedidos':           p,
            'consumos':          c,
            'devoluciones':      d,
            'inventario_fisico': inv,
            'stock_actual':      s,
        })

    return render(request, 'portal/matriz.html', {
        'usuario':    usuario,
        'filas':      filas,
        'camiones':   camiones,
        'camion_id':  camion_id or '',
        'desde':      desde,
        'hasta':      hasta,
    })


# ── Gestión de usuarios (SuperAdmin) ─────────────────────────────────────────

@_superadmin_required
def portal_usuarios(request):
    usuario = _usuario_session(request)
    empresa_id = request.GET.get('empresa')

    empresas = Empresa.objects.filter(activo=True).order_by('nombre')
    usuarios_qs = (
        Usuario.objects
        .select_related('rol', 'empresa')
        .order_by('empresa__nombre', 'nombre')
    )
    if empresa_id:
        usuarios_qs = usuarios_qs.filter(empresa_id=empresa_id)

    return render(request, 'portal/usuarios.html', {
        'usuario':    usuario,
        'usuarios':   usuarios_qs,
        'empresas':   empresas,
        'filtro_emp': empresa_id,
    })


@_superadmin_required
def portal_crear_usuario(request):
    usuario = _usuario_session(request)
    roles   = Rol.objects.all().order_by('id_rol')
    empresas = Empresa.objects.filter(activo=True).order_by('nombre')

    if request.method == 'POST':
        nombre   = request.POST.get('nombre', '').strip()
        email    = request.POST.get('email', '').strip().lower()
        clave    = request.POST.get('clave', '').strip()
        rol_id   = request.POST.get('rol_id')
        emp_id   = request.POST.get('empresa_id') or None
        telefono = request.POST.get('telefono', '').strip()

        errores = []
        if not nombre:   errores.append('El nombre es obligatorio.')
        if not email:    errores.append('El correo es obligatorio.')
        if not clave:    errores.append('La contraseña es obligatoria.')
        if not rol_id:   errores.append('Selecciona un rol.')
        if Usuario.objects.filter(email=email).exists():
            errores.append(f'El correo {email} ya está registrado.')
        try:
            rol_obj = Rol.objects.get(pk=rol_id)
        except Rol.DoesNotExist:
            errores.append('Rol inválido.')
            rol_obj = None
        if rol_obj and rol_obj.id_rol != Rol.SUPERADMIN and not emp_id:
            errores.append('Debes asignar una empresa (excepto para SuperAdmin).')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            nuevo = Usuario.objects.create(
                nombre     = nombre,
                email      = email,
                clave      = hashear_clave(clave),
                rol_id     = int(rol_id),
                empresa_id = int(emp_id) if emp_id else None,
                telefono   = telefono,
                activo     = True,
            )
            messages.success(request, f'Usuario "{nuevo.nombre}" creado correctamente.')
            return redirect('portal_usuarios')

    return render(request, 'portal/crear_usuario.html', {
        'usuario':  usuario,
        'roles':    roles,
        'empresas': empresas,
    })


@_superadmin_required
def portal_toggle_usuario(request, uid):
    if request.method != 'POST':
        return redirect('portal_usuarios')
    target = get_object_or_404(Usuario, pk=uid)
    if target.rol_id == Rol.SUPERADMIN:
        messages.error(request, 'No se puede desactivar al SuperAdmin.')
        return redirect('portal_usuarios')
    target.activo = not target.activo
    target.save(update_fields=['activo'])
    estado = 'activado' if target.activo else 'desactivado'
    messages.success(request, f'Usuario "{target.nombre}" {estado}.')
    return redirect('portal_usuarios')
