import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse

from .models import (
    Empresa, Rol, Usuario, Camion, UsuarioCamion, SST,
    Material, StockCamion, Almacen, StockAlmacen, Proveedor,
    IngresoTecsur, DetalleIngresoTecsur,
    DevolucionTecsur, DetalleDevolucionTecsur,
    MaterialMalogrado, DetalleMaterialMalogrado,
    TransferenciaAlmacen, DetalleTransferencia,
    Pedido, DetallePedido, Devolucion, DetalleDevolucion,
    UploadConsumo, Consumo, DetalleConsumo,
    Inventario, DetalleInventario,
    Suministro, SSTSuministro, SuministroTipoTrabajo,
    ManoDeObra, TipoTrabajo, SuministroManoDeObra, TipoTrabajoManoDeObra, TipoTrabajoMaterial,
    SSTEncargado, Actividad,
    Recupero, SuministroRecupero,
    ConsumoMaterialSuministro,
)


# ── Mixin Excel ───────────────────────────────────────────────────────────────
class ExcelImportMixin:
    """
    Añade dos botones al listado del admin:
      • Importar Excel  → sube un .xlsx y crea/actualiza registros
      • Descargar Plantilla → descarga un .xlsx con los encabezados correctos

    Subclases deben definir:
      excel_fields = [("campo_modelo", "Encabezado Excel"), ...]
    e implementar:
      def import_row(self, data: dict): ...
    """
    excel_fields: list = []
    change_list_template = "admin/excel_changelist.html"

    def get_urls(self):
        from django.urls import path
        mn = self.model._meta.model_name
        return [
            path("import-excel/",      self.admin_site.admin_view(self.import_excel_view),      name=f"{mn}_import_excel"),
            path("download-template/", self.admin_site.admin_view(self.download_template_view), name=f"{mn}_download_template"),
        ] + super().get_urls()

    def changelist_view(self, request, extra_context=None):
        mn  = self.model._meta.model_name
        ctx = extra_context or {}
        ctx["import_excel_url"]      = reverse(f"admin:{mn}_import_excel")
        ctx["download_template_url"] = reverse(f"admin:{mn}_download_template")
        return super().changelist_view(request, extra_context=ctx)

    # ── Descarga de plantilla ─────────────────────────────────────────────────
    def download_template_view(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"
        headers = [label for _, label in self.excel_fields]
        ws.append(headers)
        for col, cell in enumerate(ws[1], 1):
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(16, len(headers[col - 1]) + 4)
        ws.freeze_panes = "A2"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        mn = self.model._meta.model_name
        response["Content-Disposition"] = f'attachment; filename="plantilla_{mn}.xlsx"'
        wb.save(response)
        return response

    # ── Importación desde Excel ───────────────────────────────────────────────
    def import_excel_view(self, request):
        if request.method == "POST":
            archivo = request.FILES.get("archivo")
            if not archivo:
                messages.error(request, "Selecciona un archivo Excel (.xlsx).")
                return redirect(".")
            try:
                wb   = openpyxl.load_workbook(archivo, data_only=True)
                rows = list(wb.active.iter_rows(values_only=True))
            except Exception as exc:
                messages.error(request, f"Error al leer el archivo: {exc}")
                return redirect(".")
            if len(rows) < 2:
                messages.error(request, "El archivo no contiene datos (solo encabezado).")
                return redirect(".")

            creados = errores = 0
            for num, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue
                data = {
                    field: (row[i] if i < len(row) else None)
                    for i, (field, _) in enumerate(self.excel_fields)
                }
                try:
                    self.import_row(data)
                    creados += 1
                except Exception as exc:
                    messages.warning(request, f"Fila {num}: {exc}")
                    errores += 1

            if creados:
                messages.success(request, f"{creados} registro(s) importado(s) correctamente.")
            if errores:
                messages.error(request, f"{errores} fila(s) con error — revisa las advertencias.")
            return redirect("..")

        context = {
            **self.admin_site.each_context(request),
            "title":  f"Importar {self.model._meta.verbose_name_plural}",
            "opts":   self.model._meta,
            "fields": self.excel_fields,
        }
        return render(request, "admin/excel_import_form.html", context)

    def import_row(self, data: dict):
        raise NotImplementedError("Implementa import_row() en tu ModelAdmin.")


# ── Helpers ───────────────────────────────────────────────────────────────────
def _str(val, default=""):
    return str(val).strip() if val is not None else default

def _hash_clave(clave: str) -> str:
    """PBKDF2 seguro. Si ya viene hasheado (pbkdf2_ o sha256 legacy) lo deja igual."""
    from .security import hashear_clave
    h = str(clave).strip()
    # Ya hasheado en formato PBKDF2
    if h.startswith('pbkdf2_') or h.startswith('bcrypt'):
        return h
    # Legacy SHA256 (64 hex) — se migra al primer login, no re-hashear
    if len(h) == 64 and all(c in '0123456789abcdef' for c in h):
        return h
    return hashear_clave(h)


# ── Empresa ───────────────────────────────────────────────────────────────────
@admin.register(Empresa)
class EmpresaAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["nombre", "ruc", "activo"]
    search_fields = ["nombre", "ruc"]
    excel_fields  = [
        ("nombre",    "Nombre"),
        ("ruc",       "RUC"),
        ("direccion", "Dirección"),
        ("telefono",  "Teléfono"),
        ("email",     "Email"),
    ]

    def import_row(self, data):
        Empresa.objects.update_or_create(
            ruc=_str(data["ruc"]),
            defaults={
                "nombre":    _str(data["nombre"]),
                "direccion": _str(data.get("direccion")),
                "telefono":  _str(data.get("telefono")),
                "email":     _str(data.get("email")),
            },
        )


# ── Rol ───────────────────────────────────────────────────────────────────────
@admin.register(Rol)
class RolAdmin(admin.ModelAdmin):
    list_display = ["id_rol", "descripcion"]


# ── Usuario ───────────────────────────────────────────────────────────────────
@admin.register(Usuario)
class UsuarioAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["nombre", "email", "rol", "empresa", "activo"]
    list_filter   = ["rol", "activo", "empresa"]
    search_fields = ["nombre", "email"]
    excel_fields  = [
        ("nombre",   "Nombre completo"),
        ("email",    "Email"),
        ("clave",    "Contraseña"),
        ("rol",      "Rol (ej: Capataz)"),
        ("empresa",  "Empresa (nombre)"),
        ("telefono", "Teléfono"),
    ]

    def save_model(self, request, obj, form, change):
        obj.clave = _hash_clave(obj.clave)
        super().save_model(request, obj, form, change)

    def import_row(self, data):
        rol = Rol.objects.get(descripcion__icontains=_str(data["rol"]))
        empresa = None
        if data.get("empresa") and _str(data["empresa"]):
            empresa = Empresa.objects.get(nombre__icontains=_str(data["empresa"]))
        Usuario.objects.update_or_create(
            email=_str(data["email"]),
            defaults={
                "nombre":   _str(data["nombre"]),
                "clave":    _hash_clave(_str(data["clave"])),
                "rol":      rol,
                "empresa":  empresa,
                "telefono": _str(data.get("telefono")),
            },
        )


# ── Camion ────────────────────────────────────────────────────────────────────
@admin.register(Camion)
class CamionAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display = ["placa", "empresa", "activo"]
    excel_fields = [
        ("empresa",     "Empresa (nombre)"),
        ("placa",       "Placa"),
        ("descripcion", "Descripción"),
    ]

    def import_row(self, data):
        empresa = Empresa.objects.get(nombre__icontains=_str(data["empresa"]))
        Camion.objects.update_or_create(
            placa=_str(data["placa"]),
            defaults={
                "empresa":     empresa,
                "descripcion": _str(data.get("descripcion")),
            },
        )


# ── UsuarioCamion ─────────────────────────────────────────────────────────────
@admin.register(UsuarioCamion)
class UsuarioCamionAdmin(admin.ModelAdmin):
    list_display = ["usuario", "camion", "fecha_inicio", "fecha_fin", "activo"]


# ── SST ───────────────────────────────────────────────────────────────────────
class SSTSuministroInline(admin.TabularInline):
    model  = SSTSuministro
    extra  = 0
    fields = ["suministro", "asignado_a"]
    autocomplete_fields = ["suministro"]

@admin.register(SSTSuministro)
class SSTSuministroAdmin(admin.ModelAdmin):
    list_display        = ["sst", "suministro", "asignado_a"]
    search_fields       = ["sst__codigo", "suministro__numero_suministro"]
    list_select_related = True

class SSTEncargadoInline(admin.TabularInline):
    model  = SSTEncargado
    extra  = 0
    fields = ["usuario"]

@admin.register(SSTEncargado)
class SSTEncargadoAdmin(admin.ModelAdmin):
    list_display        = ["sst", "usuario"]
    search_fields       = ["sst__codigo", "usuario__nombre"]
    autocomplete_fields = ["usuario"]

@admin.register(SST)
class SSTAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["codigo", "distrito", "actividad", "empresa", "fecha_inicio", "fecha_termino", "monto_sst"]
    search_fields = ["codigo", "distrito"]
    list_filter   = ["empresa"]
    inlines       = [SSTEncargadoInline, SSTSuministroInline]
    excel_fields  = [
        ("empresa",       "Empresa (nombre)"),
        ("codigo",        "Código SST"),
        ("distrito",      "Distrito"),
        ("actividad",     "Actividad"),
        ("fecha_inicio",  "Fecha Inicio"),
        ("fecha_termino", "Fecha Término"),
        ("monto_sst",     "Monto SST"),
    ]

    def import_row(self, data):
        empresa = Empresa.objects.get(nombre__icontains=_str(data["empresa"]))
        actividad = None
        if _str(data.get("actividad")):
            actividad, _ = Actividad.objects.get_or_create(nombre=_str(data["actividad"]))
        SST.objects.update_or_create(
            codigo=_str(data["codigo"]),
            empresa=empresa,
            defaults={
                "distrito":      _str(data["distrito"]),
                "actividad":     actividad,
                "fecha_inicio":  data.get("fecha_inicio") or None,
                "fecha_termino": data.get("fecha_termino") or None,
                "monto_sst":     data.get("monto_sst") or 0,
            },
        )


# ── Suministro ────────────────────────────────────────────────────────────────
class SuministroTipoTrabajoInline(admin.TabularInline):
    model  = SuministroTipoTrabajo
    extra  = 0
    fields = ["tipo_trabajo"]

class SuministroManoDeObraInline(admin.TabularInline):
    model  = SuministroManoDeObra
    extra  = 0
    fields = ["mano_de_obra"]

class SuministroRecuperoInline(admin.TabularInline):
    model  = SuministroRecupero
    extra  = 0
    fields = ["recupero", "cantidad", "fecha"]

@admin.register(SuministroTipoTrabajo)
class SuministroTipoTrabajoAdmin(admin.ModelAdmin):
    list_display        = ["suministro", "tipo_trabajo"]
    search_fields       = ["suministro__numero_suministro", "tipo_trabajo__nombre"]
    list_select_related = True

@admin.register(SuministroManoDeObra)
class SuministroManoDeObraAdmin(admin.ModelAdmin):
    list_display        = ["suministro", "mano_de_obra"]
    search_fields       = ["suministro__numero_suministro", "mano_de_obra__partida"]
    list_select_related = True

@admin.register(Suministro)
class SuministroAdmin(ExcelImportMixin, admin.ModelAdmin):
    inlines = [SuministroTipoTrabajoInline, SuministroManoDeObraInline, SuministroRecuperoInline]
    list_display  = ["numero_suministro", "medidor", "distrito", "estado", "monto_sum", "fecha_ejecucion"]
    list_filter   = ["estado"]
    search_fields = ["numero_suministro", "medidor"]
    excel_fields  = [
        ("numero_suministro", "N° Suministro"),
        ("medidor",           "Medidor"),
        ("distrito",          "Distrito"),
        ("monto_sum",         "Monto (S/)"),
        ("estado",            "Estado"),
        ("ejecutado_por",     "Ejecutado Por"),
    ]

    def import_row(self, data):
        Suministro.objects.update_or_create(
            numero_suministro=_str(data["numero_suministro"]),
            defaults={
                "medidor":      _str(data.get("medidor")),
                "distrito":     _str(data.get("distrito")),
                "monto_sum":    data.get("monto_sum") or 0,
                "estado":       _str(data.get("estado")) or "asignado",
                "ejecutado_por": _str(data.get("ejecutado_por")),
            },
        )


# ── ManoDeObra ────────────────────────────────────────────────────────────────
@admin.register(ManoDeObra)
class ManoDeObraAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["partida", "descripcion", "precio"]
    search_fields = ["partida", "descripcion"]
    excel_fields  = [
        ("partida",     "Partida"),
        ("descripcion", "Descripción"),
        ("precio",      "Precio"),
    ]

    def import_row(self, data):
        ManoDeObra.objects.update_or_create(
            partida=_str(data["partida"]),
            defaults={
                "descripcion": _str(data["descripcion"]),
                "precio":      data["precio"],
            },
        )


# ── Material ──────────────────────────────────────────────────────────────────
@admin.register(Material)
class MaterialAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["matricula", "descripcion", "precio"]
    search_fields = ["matricula", "descripcion"]
    excel_fields  = [
        ("matricula",   "Matrícula"),
        ("descripcion", "Descripción"),
        ("precio",      "Precio"),
    ]

    def import_row(self, data):
        Material.objects.update_or_create(
            matricula=_str(data["matricula"]),
            defaults={
                "descripcion": _str(data["descripcion"]),
                "precio":      data["precio"],
            },
        )


# ── StockCamion ───────────────────────────────────────────────────────────────
@admin.register(StockCamion)
class StockCamionAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display = ["camion", "material", "cantidad"]
    excel_fields = [
        ("camion",   "Camión (placa)"),
        ("material", "Material (matrícula)"),
        ("cantidad", "Cantidad"),
    ]

    def import_row(self, data):
        camion   = Camion.objects.get(placa=_str(data["camion"]))
        material = Material.objects.get(matricula=_str(data["material"]))
        StockCamion.objects.update_or_create(
            camion=camion, material=material,
            defaults={"cantidad": int(data["cantidad"])},
        )


# ── Almacen ───────────────────────────────────────────────────────────────────
@admin.register(Almacen)
class AlmacenAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display = ["nombre", "empresa", "activo"]
    excel_fields = [
        ("empresa",   "Empresa (nombre)"),
        ("nombre",    "Nombre del almacén"),
        ("direccion", "Dirección"),
    ]

    def import_row(self, data):
        empresa = Empresa.objects.get(nombre__icontains=_str(data["empresa"]))
        Almacen.objects.update_or_create(
            nombre=_str(data["nombre"]), empresa=empresa,
            defaults={"direccion": _str(data.get("direccion"))},
        )


# ── StockAlmacen ──────────────────────────────────────────────────────────────
@admin.register(StockAlmacen)
class StockAlmacenAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display = ["almacen", "material", "cantidad"]
    excel_fields = [
        ("almacen",  "Almacén (nombre)"),
        ("material", "Material (matrícula)"),
        ("cantidad", "Cantidad"),
    ]

    def import_row(self, data):
        almacen  = Almacen.objects.get(nombre__icontains=_str(data["almacen"]))
        material = Material.objects.get(matricula=_str(data["material"]))
        StockAlmacen.objects.update_or_create(
            almacen=almacen, material=material,
            defaults={"cantidad": int(data["cantidad"])},
        )


# ── Proveedor ─────────────────────────────────────────────────────────────────
@admin.register(Proveedor)
class ProveedorAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["nombre", "ruc", "activo"]
    search_fields = ["nombre", "ruc"]
    excel_fields  = [
        ("nombre",   "Nombre"),
        ("ruc",      "RUC"),
        ("direccion","Dirección"),
        ("telefono", "Teléfono"),
        ("email",    "Email"),
        ("contacto", "Contacto"),
    ]

    def import_row(self, data):
        Proveedor.objects.update_or_create(
            ruc=_str(data["ruc"]),
            defaults={
                "nombre":    _str(data["nombre"]),
                "direccion": _str(data.get("direccion")),
                "telefono":  _str(data.get("telefono")),
                "email":     _str(data.get("email")),
                "contacto":  _str(data.get("contacto")),
            },
        )


# ── IngresoTecsur ─────────────────────────────────────────────────────────────
class DetalleIngresoInline(admin.TabularInline):
    model = DetalleIngresoTecsur
    extra = 0

@admin.register(IngresoTecsur)
class IngresoTecsurAdmin(admin.ModelAdmin):
    list_display = ["folio", "almacen", "proveedor", "fecha"]
    inlines      = [DetalleIngresoInline]


# ── DevolucionTecsur ──────────────────────────────────────────────────────────
class DetalleDevTecsurInline(admin.TabularInline):
    model = DetalleDevolucionTecsur
    extra = 0

@admin.register(DevolucionTecsur)
class DevolucionTecsurAdmin(admin.ModelAdmin):
    list_display = ["folio", "almacen", "proveedor", "fecha"]
    inlines      = [DetalleDevTecsurInline]


# ── MaterialMalogrado ─────────────────────────────────────────────────────────
class DetalleMalogradoInline(admin.TabularInline):
    model = DetalleMaterialMalogrado
    extra = 0

@admin.register(MaterialMalogrado)
class MaterialMalogradoAdmin(admin.ModelAdmin):
    list_display = ["folio_factura", "almacen", "fecha"]
    inlines      = [DetalleMalogradoInline]


# ── TransferenciaAlmacen ──────────────────────────────────────────────────────
class DetalleTransferenciaInline(admin.TabularInline):
    model = DetalleTransferencia
    extra = 0

@admin.register(TransferenciaAlmacen)
class TransferenciaAlmacenAdmin(admin.ModelAdmin):
    list_display = ["almacen_origen", "almacen_destino", "fecha"]
    inlines      = [DetalleTransferenciaInline]


# ── Pedido ────────────────────────────────────────────────────────────────────
class DetallePedidoInline(admin.TabularInline):
    model = DetallePedido
    extra = 0

@admin.register(Pedido)
class PedidoAdmin(admin.ModelAdmin):
    list_display = ["id_pedido", "camion", "usuario", "estado", "fecha"]
    list_filter  = ["estado"]
    inlines      = [DetallePedidoInline]


# ── Devolucion ────────────────────────────────────────────────────────────────
class DetalleDevolucionInline(admin.TabularInline):
    model = DetalleDevolucion
    extra = 0

@admin.register(Devolucion)
class DevolucionAdmin(admin.ModelAdmin):
    list_display = ["id_devolucion", "camion", "usuario", "estado", "fecha"]
    list_filter  = ["estado"]
    inlines      = [DetalleDevolucionInline]


# ── UploadConsumo ─────────────────────────────────────────────────────────────
@admin.register(UploadConsumo)
class UploadConsumoAdmin(admin.ModelAdmin):
    list_display = ["sst", "usuario", "estado", "fecha_upload"]
    list_filter  = ["estado"]


# ── Inventario ────────────────────────────────────────────────────────────────
@admin.register(Inventario)
class InventarioAdmin(admin.ModelAdmin):
    list_display = ["id_inventario", "camion", "almacen", "mes", "anio", "estado"]
    list_filter  = ["estado"]


# ── Actividad ─────────────────────────────────────────────────────────────────
@admin.register(Actividad)
class ActividadAdmin(admin.ModelAdmin):
    list_display  = ["nombre"]
    search_fields = ["nombre"]


# ── TipoTrabajo ───────────────────────────────────────────────────────────────
class TipoTrabajoManoDeObraInline(admin.TabularInline):
    model  = TipoTrabajoManoDeObra
    extra  = 1
    fields = ["mano_de_obra"]
    autocomplete_fields = ["mano_de_obra"]


class TipoTrabajoMaterialInline(admin.TabularInline):
    model  = TipoTrabajoMaterial
    extra  = 1
    fields = ["material", "cantidad"]
    autocomplete_fields = ["material"]


@admin.register(TipoTrabajo)
class TipoTrabajoAdmin(admin.ModelAdmin):
    inlines       = [TipoTrabajoManoDeObraInline, TipoTrabajoMaterialInline]
    list_display  = ["nombre"]
    search_fields = ["nombre"]


# ── Recupero ──────────────────────────────────────────────────────────────────
@admin.register(Recupero)
class RecuperoAdmin(ExcelImportMixin, admin.ModelAdmin):
    list_display  = ["matricula", "descripcion"]
    search_fields = ["matricula", "descripcion"]
    excel_fields  = [
        ("matricula",   "Matrícula"),
        ("descripcion", "Descripción"),
    ]

    def import_row(self, data):
        Recupero.objects.update_or_create(
            matricula=_str(data["matricula"]),
            defaults={"descripcion": _str(data["descripcion"])},
        )


# ── SuministroRecupero ────────────────────────────────────────────────────────
@admin.register(SuministroRecupero)
class SuministroRecuperoAdmin(admin.ModelAdmin):
    list_display        = ["suministro", "recupero", "cantidad", "fecha"]
    search_fields       = ["suministro__numero_suministro", "recupero__matricula"]
    list_select_related = True


# ── ConsumoMaterialSuministro ─────────────────────────────────────────────────
@admin.register(ConsumoMaterialSuministro)
class ConsumoMaterialSuministroAdmin(admin.ModelAdmin):
    list_display        = ["liquidacion", "suministro", "material", "cantidad", "usuario", "fecha"]
    list_filter         = ["fecha"]
    search_fields       = ["suministro__numero_suministro", "material__matricula", "material__descripcion"]
    list_select_related = True
